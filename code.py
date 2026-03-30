import os
import csv
import secrets
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIG ──────────────────────────────────────────────────────────────────
OUTPUT_DIR = r"C:\Users\hamza.zahid\OneDrive - mgapparel.com\Desktop\Budget Mails\Generic"
OUTPUT_FILE = "Budget_FY2027_MGA.xlsx"  # Generate valid .xlsx; save as .xlsm after importing VBA

PASSWORD_CSV_PATH = Path(__file__).with_name("DepartmentPasswords_CONFIDENTIAL.csv")

DEPARTMENTS = [
    "ACCOUNTS",
    "AUDIT",
    "BUSINESS AFFAIRS, SUSTAINABILITY, CSR",
    "PD & SAMPLING",
    "MARKETING & MERCHANDIZING",
    "RESEARCH & DESIGN",
    "FABRIC SOURCING",
    "EXPORT & LOGISTICS",
    "MATERIAL MANAGEMENT & CONTROL",
    "ENGINEERING & UTILITIES",
    "ADMINISTRATION",
    "COMPLIANCE, HSE & IR",
    "STORES",
    "HUMAN RESOURCES",
    "CUTTING & EMBROIDERY",
    "STITCHING",
    "WASHING & DRY PROCESS",
    "FINISHING",
    "PPC & WIP",
    "IE & PROCESS IMPROVEMENT",
    "MAINTENANCE",
    "MIS & IT",
    "QUALITY ASSURANCE",
    "QUALITY CONTROL",
    "R61 OPERATIONS",
    "DSBA",
]


def _random_password(length: int = 18) -> str:
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnopqrstuvwxyz23456789"
    return "".join(secrets.choice(alphabet) for _ in range(length))


def _write_password_csv(master_password: str, dept_passwords: dict) -> None:
    PASSWORD_CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(PASSWORD_CSV_PATH, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Department", "Password"])
        w.writerow(["[MASTER]", master_password])
        for dept in DEPARTMENTS:
            w.writerow([dept, dept_passwords[dept]])


def load_passwords() -> tuple[str, dict]:
    if not PASSWORD_CSV_PATH.exists():
        master_password = _random_password(24)
        dept_passwords = {dept: _random_password(20) for dept in DEPARTMENTS}
        _write_password_csv(master_password, dept_passwords)
        print("Created DepartmentPasswords_CONFIDENTIAL.csv with random passwords:")
        print(f"  {PASSWORD_CSV_PATH}")
        return master_password, dept_passwords

    dept_passwords: dict[str, str] = {}
    master_password = ""
    with open(PASSWORD_CSV_PATH, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            dept = (row.get("Department") or "").strip().strip("\ufeff")
            pw = (row.get("Password") or "").strip()
            if not dept or not pw:
                continue
            if dept == "[MASTER]":
                master_password = pw
            else:
                dept_passwords[dept.strip('"')] = pw

    missing = [d for d in DEPARTMENTS if d not in dept_passwords]
    if missing:
        raise RuntimeError(
            "Missing passwords for departments in DepartmentPasswords_CONFIDENTIAL.csv: "
            + ", ".join(missing)
        )
    if not master_password:
        raise RuntimeError("Missing [MASTER] password in DepartmentPasswords_CONFIDENTIAL.csv")

    return master_password, dept_passwords


MASTER_PASSWORD, DEPT_PASSWORDS = load_passwords()

MONTHS = ["Jul-26","Aug-26","Sep-26","Oct-26","Nov-26","Dec-26",
          "Jan-27","Feb-27","Mar-27","Apr-27","May-27","Jun-27"]

SECTIONS = [
    ("1. TRAINING & DEVELOPMENT", [
        ("Employee 1",  [("Units/Persons","units"),("Rate (PKR)","rate"),("Amount (PKR)","value")]),
        ("Employee 2",  [("Units/Persons","units"),("Rate (PKR)","rate"),("Amount (PKR)","value")]),
        ("Employee 3",  [("Units/Persons","units"),("Rate (PKR)","rate"),("Amount (PKR)","value")]),
    ]),
    ("2. SOFTWARE & LICENSES", [
        ("Platform A",  [("Units/Licenses","units"),("Rate (PKR)","rate"),("Amount (PKR)","value")]),
        ("Platform B",  [("Units/Licenses","units"),("Rate (PKR)","rate"),("Amount (PKR)","value")]),
        ("Platform C",  [("Units/Licenses","units"),("Rate (PKR)","rate"),("Amount (PKR)","value")]),
    ]),
    ("3. IT EQUIPMENT (CAPEX)", [
        ("Laptops",     [("Units/Pcs","units"),("Rate (PKR)","rate"),("Amount (PKR)","value")]),
        ("Others",      [("Units/Pcs","units"),("Rate (PKR)","rate"),("Amount (PKR)","value")]),
    ]),
    ("4. STORE CONSUMPTION", [
        ("Pens",        [("Units/Pcs","units"),("Rate (PKR)","rate"),("Amount (PKR)","value")]),
        ("Paper (Ream)",[("Units/Pcs","units"),("Rate (PKR)","rate"),("Amount (PKR)","value")]),
        ("Toner/Ink",   [("Units/Pcs","units"),("Rate (PKR)","rate"),("Amount (PKR)","value")]),
        ("Other Items", [("Units/Pcs","units"),("Rate (PKR)","rate"),("Amount (PKR)","value")]),
    ]),
    ("5. ENTERTAINMENT", [
        ("Client Entertainment",[("Occasions","units"),("Rate (PKR)","rate"),("Amount (PKR)","value")]),
        ("Staff Events",        [("Events","units"),   ("Rate (PKR)","rate"),("Amount (PKR)","value")]),
    ]),
    ("6. FOREIGN TRAVEL (USD)", [
        ("Airfare",       [("Trips","units"),    ("Rate (USD)","rate"),  ("Amount (USD)","value")]),
        ("Food & Lodging",[("Days/Stay","units"),("Allowance (USD)","rate"),("Amount (USD)","value")]),
        ("Other Costs",   [("Lump Sum (USD)","other")]),
    ]),
    ("7. LOCAL TRAVEL (PKR)", [
        ("Multan",    [("Trips","units"),("Rate (PKR)","rate"),("Amount (PKR)","value")]),
        ("Lahore",    [("Trips","units"),("Rate (PKR)","rate"),("Amount (PKR)","value")]),
        ("Islamabad", [("Trips","units"),("Rate (PKR)","rate"),("Amount (PKR)","value")]),
        ("Karachi",   [("Trips","units"),("Rate (PKR)","rate"),("Amount (PKR)","value")]),
        ("Other",     [("Trips","units"),("Rate (PKR)","rate"),("Amount (PKR)","value")]),
    ]),
]

# ─── COLOURS ─────────────────────────────────────────────────────────────────
DARK_NAVY   = "1F3864"
MID_BLUE    = "2E5FAC"
ACCENT_TEAL = "1ABC9C"
HEADER_GOLD = "F0C040"
SECTION_BG  = "D6E4F0"
INPUT_YELLOW= "FFFACD"
FORMULA_BG  = "E8F8F5"
TOTAL_BG    = "D5F5E3"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"
ALT_ROW     = "F4F9FD"
BORDER_CLR  = "BFBFBF"

def side(style="thin"):
    return Side(style=style, color=BORDER_CLR)

def thin():
    s = side()
    return Border(left=s, right=s, top=s, bottom=s)

def thick():
    s = side("medium")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(cell, text, bg=DARK_NAVY, fg=WHITE, size=10, bold=True, align="center", wrap=False):
    cell.value = text
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = thin()

def sec(cell, text):
    cell.value = text
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=MID_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = thin()

def lbl(cell, text, bold=False, bg=None, indent=0):
    cell.value = text
    cell.font = Font(name="Arial", bold=bold, color="222222", size=9)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
    cell.border = thin()

def inp(cell):
    cell.font = Font(name="Arial", color="00008B", size=9)
    cell.fill = PatternFill("solid", fgColor=INPUT_YELLOW)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin()
    cell.number_format = '#,##0;(#,##0);"-"'

def frm(cell, formula):
    cell.value = formula
    cell.font = Font(name="Arial", color="000000", size=9)
    cell.fill = PatternFill("solid", fgColor=FORMULA_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin()
    cell.number_format = '#,##0;(#,##0);"-"'

def tot(cell, formula, bg=TOTAL_BG, bold=True):
    cell.value = formula
    cell.font = Font(name="Arial", bold=bold, color="1A5276", size=9)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thick()
    cell.number_format = '#,##0;(#,##0);"-"'

def blank(ws, row, c1, c2, bg=WHITE):
    for c in range(c1, c2+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = thin()

# ─── BUILD DEPT SHEET ─────────────────────────────────────────────────────────
def build_dept(wb, dept_name):
    # Sheet name limited to 31 chars
    sname = dept_name[:31]
    ws = wb.create_sheet(sname)
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    for i in range(4, 16):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.column_dimensions[get_column_letter(16)].width = 12
    ws.column_dimensions[get_column_letter(17)].width = 22

    # Row 1 – Banner
    ws.row_dimensions[1].height = 34
    ws.merge_cells("A1:Q1")
    c = ws["A1"]
    c.value = f"ANNUAL BUDGET FY 2026-27  |  {dept_name}"
    c.font = Font(name="Arial", bold=True, color=WHITE, size=13)
    c.fill = PatternFill("solid", fgColor=DARK_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2 – Instruction bar
    ws.row_dimensions[2].height = 18
    ws.merge_cells("A2:Q2")
    c = ws["A2"]
    c.value = "Fill YELLOW cells only (Units, Rate, Trips etc). Blue cells are auto-calculated. Add assumptions in the Comments column."
    c.font = Font(name="Arial", italic=True, color="7B241C", size=8)
    c.fill = PatternFill("solid", fgColor="FDEBD0")
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 3 – Column headers
    ws.row_dimensions[3].height = 26
    hdr(ws["A3"], "#")
    hdr(ws["B3"], "Expense Category", align="left")
    hdr(ws["C3"], "Sub-Type / Unit", wrap=True)
    for i, m in enumerate(MONTHS, start=4):
        hdr(ws.cell(row=3, column=i), m, bg=MID_BLUE, size=9)
    hdr(ws.cell(row=3, column=16), "FY Total", bg=HEADER_GOLD, fg=DARK_NAVY)
    hdr(ws.cell(row=3, column=17), "Comments / Assumptions", wrap=True)

    row = 4
    section_total_rows = {}

    for sec_title, items in SECTIONS:
        ws.row_dimensions[row].height = 20
        ws.merge_cells(f"A{row}:Q{row}")
        sec(ws[f"A{row}"], f"  {sec_title}")
        row += 1

        value_rows_in_section = []

        for item_name, sub_rows in items:
            units_row = rate_row = None
            for sub_label, rtype in sub_rows:
                ws.row_dimensions[row].height = 17
                bg = ALT_ROW if row % 2 == 0 else WHITE

                # Col A
                ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=SECTION_BG)
                ws.cell(row=row, column=1).border = thin()

                # Col B
                b_cell = ws.cell(row=row, column=2)
                if rtype == "units":
                    lbl(b_cell, item_name, bold=True, bg=SECTION_BG, indent=1)
                else:
                    lbl(b_cell, "", bg=bg)

                # Col C
                lbl(ws.cell(row=row, column=3), sub_label, bg=bg, indent=1)

                if rtype == "units":
                    units_row = row
                    for col in range(4, 16):
                        inp(ws.cell(row=row, column=col))
                    frm(ws.cell(row=row, column=16),
                        f"=SUM({get_column_letter(4)}{row}:{get_column_letter(15)}{row})")
                    ws.cell(row=row, column=17).border = thin()

                elif rtype == "rate":
                    rate_row = row
                    for col in range(4, 16):
                        inp(ws.cell(row=row, column=col))
                    frm(ws.cell(row=row, column=16),
                        f'=IFERROR(AVERAGEIF({get_column_letter(4)}{row}:{get_column_letter(15)}{row},"<>0"),"-")')
                    ws.cell(row=row, column=17).border = thin()

                elif rtype == "value":
                    if units_row and rate_row:
                        for col in range(4, 16):
                            cl = get_column_letter(col)
                            frm(ws.cell(row=row, column=col),
                                f"={cl}{units_row}*{cl}{rate_row}")
                        frm(ws.cell(row=row, column=16),
                            f"=SUM({get_column_letter(4)}{row}:{get_column_letter(15)}{row})")
                        value_rows_in_section.append(row)
                    ws.cell(row=row, column=17).border = thin()

                elif rtype == "other":
                    for col in range(4, 16):
                        inp(ws.cell(row=row, column=col))
                    frm(ws.cell(row=row, column=16),
                        f"=SUM({get_column_letter(4)}{row}:{get_column_letter(15)}{row})")
                    value_rows_in_section.append(row)
                    ws.cell(row=row, column=17).border = thin()

                row += 1

            blank(ws, row, 1, 17)
            row += 1

        # Section total row
        ws.row_dimensions[row].height = 20
        ws.merge_cells(f"A{row}:C{row}")
        c = ws[f"A{row}"]
        sec_short = sec_title.split(". ", 1)[1] if ". " in sec_title else sec_title
        c.value = f"  {sec_short}  —  TOTAL"
        c.font = Font(name="Arial", bold=True, color=WHITE, size=9)
        c.fill = PatternFill("solid", fgColor=ACCENT_TEAL)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thick()
        ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=ACCENT_TEAL)
        ws.cell(row=row, column=3).border = thick()

        for col in range(4, 17):
            cl = get_column_letter(col)
            if value_rows_in_section:
                formula = "=" + "+".join([f"{cl}{vr}" for vr in value_rows_in_section])
            else:
                formula = "=0"
            tot(ws.cell(row=row, column=col), formula, bg=TOTAL_BG)
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=ACCENT_TEAL)
            ws.cell(row=row, column=col).font = Font(name="Arial", bold=True, color=WHITE, size=9)
            ws.cell(row=row, column=col).border = thick()

        ws.cell(row=row, column=17).fill = PatternFill("solid", fgColor=ACCENT_TEAL)
        ws.cell(row=row, column=17).border = thick()

        section_total_rows[sec_title] = row
        row += 2

    # Grand total
    ws.row_dimensions[row].height = 24
    ws.merge_cells(f"A{row}:C{row}")
    c = ws[f"A{row}"]
    c.value = "  DEPARTMENT GRAND TOTAL"
    c.font = Font(name="Arial", bold=True, color=WHITE, size=11)
    c.fill = PatternFill("solid", fgColor=DARK_NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = thick()
    ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=DARK_NAVY)
    ws.cell(row=row, column=3).border = thick()

    for col in range(4, 17):
        cl = get_column_letter(col)
        formula = "=" + "+".join([f"{cl}{r}" for r in section_total_rows.values()])
        c2 = ws.cell(row=row, column=col)
        c2.value = formula
        c2.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        c2.fill = PatternFill("solid", fgColor=DARK_NAVY)
        c2.border = thick()
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.number_format = '#,##0;(#,##0);"-"'

    ws.cell(row=row, column=17).fill = PatternFill("solid", fgColor=DARK_NAVY)
    ws.cell(row=row, column=17).border = thick()

    ws.freeze_panes = "D4"
    return section_total_rows, row


# ─── BUILD WORKBOOK ──────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# Instructions sheet (always visible)
ws_inst = wb.create_sheet("INSTRUCTIONS", 0)
ws_inst.sheet_view.showGridLines = False
ws_inst.column_dimensions["A"].width = 4
ws_inst.column_dimensions["B"].width = 28
ws_inst.column_dimensions["C"].width = 55

inst_rows = [
    ("ANNUAL BUDGET FY 2026-27 — HOW TO USE", None, None, DARK_NAVY, WHITE, 13, True, 32),
    ("STEP", "ACTION", "DETAIL", MID_BLUE, WHITE, 10, True, 22),
    ("1", "Receive this file via OneDrive link", "The file is shared with you. Do not download — open directly in OneDrive.", INPUT_YELLOW, "222222", 9, False, 20),
    ("2", "Enter your department password when prompted", "A password box appears when you open the file. Enter the password shared with you by Finance.", FORMULA_BG, "222222", 9, False, 20),
    ("3", "Your department tab will appear", "Only your department's sheet will be visible and unlocked.", TOTAL_BG, "222222", 9, False, 20),
    ("4", "Fill YELLOW cells only", "Enter Units, Rate, Trips, Days etc. in yellow cells. DO NOT edit blue cells — they are auto-calculated.", INPUT_YELLOW, "222222", 9, False, 20),
    ("5", "Add comments in the last column", "Note vendor names, assumptions, or anything Finance should know.", FORMULA_BG, "222222", 9, False, 20),
    ("6", "Save the file (Ctrl+S)", "Your data is saved in the shared file on OneDrive. No need to email anything back.", TOTAL_BG, "222222", 9, False, 20),
    (None, None, None, WHITE, WHITE, 9, False, 8),
    ("COLOUR LEGEND", None, None, DARK_NAVY, WHITE, 11, True, 22),
    ("", "Yellow cells", "INPUT — Fill these with your figures", INPUT_YELLOW, "222222", 9, False, 18),
    ("", "Blue text cells", "FORMULA — Auto-calculated, do not edit", FORMULA_BG, "222222", 9, False, 18),
    ("", "Teal cells", "Section totals — sum of all items in that category", TOTAL_BG, "222222", 9, False, 18),
    ("", "Dark navy cells", "Department Grand Total", LIGHT_GRAY, "222222", 9, False, 18),
    (None, None, None, WHITE, WHITE, 9, False, 8),
    ("IMPORTANT", None, None, "C0392B", WHITE, 11, True, 22),
    ("!", "Do NOT add or delete rows", "This will break the Consolidated Summary formulas.", "FDEBD0", "7B241C", 9, False, 18),
    ("!", "Foreign Travel (Section 6)", "Amounts are in USD. All other sections are in PKR.", "FDEBD0", "7B241C", 9, False, 18),
    ("!", "Do NOT share your password", "Your password is unique to your department. Keep it confidential.", "FDEBD0", "7B241C", 9, False, 18),
    ("!", "Deadline", "All data must be entered by [DATE]. Contact Finance for extensions.", "FDEBD0", "7B241C", 9, False, 18),
]

for ridx, rdata in enumerate(inst_rows, start=1):
    ws_inst.row_dimensions[ridx].height = rdata[7]
    if rdata[0] is None:
        for c in range(1, 4):
            ws_inst.cell(row=ridx, column=c).fill = PatternFill("solid", fgColor=rdata[3])
        continue
    if rdata[1] is None:
        ws_inst.merge_cells(f"A{ridx}:C{ridx}")
        c = ws_inst[f"A{ridx}"]
        c.value = rdata[0]
        c.font = Font(name="Arial", bold=rdata[6], color=rdata[4], size=rdata[5])
        c.fill = PatternFill("solid", fgColor=rdata[3])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()
    else:
        for cidx, val in enumerate([rdata[0], rdata[1], rdata[2]], start=1):
            c = ws_inst.cell(row=ridx, column=cidx)
            c.value = val
            c.font = Font(name="Arial", bold=(rdata[6] and cidx <= 2), color=rdata[4], size=rdata[5])
            c.fill = PatternFill("solid", fgColor=rdata[3])
            c.alignment = Alignment(horizontal="left" if cidx > 1 else "center",
                                    vertical="center", wrap_text=True)
            c.border = thin()

# Summary sheet (always visible)
ws_sum = wb.create_sheet("CONSOLIDATED SUMMARY", 1)
ws_sum.sheet_view.showGridLines = False
ws_sum.sheet_properties.tabColor = DARK_NAVY

end_col = 2 + len(DEPARTMENTS) + 1
ws_sum.column_dimensions["A"].width = 4
ws_sum.column_dimensions["B"].width = 36
for i in range(3, end_col + 1):
    ws_sum.column_dimensions[get_column_letter(i)].width = 18
ws_sum.column_dimensions[get_column_letter(end_col)].width = 18

ws_sum.row_dimensions[1].height = 36
ws_sum.merge_cells(f"A1:{get_column_letter(end_col)}1")
c = ws_sum["A1"]
c.value = "ANNUAL BUDGET FY 2026-27 — CONSOLIDATED SUMMARY"
c.font = Font(name="Arial", bold=True, color=WHITE, size=14)
c.fill = PatternFill("solid", fgColor=DARK_NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")

ws_sum.row_dimensions[2].height = 18
ws_sum.merge_cells(f"A2:{get_column_letter(end_col)}2")
c = ws_sum["A2"]
c.value = "This sheet auto-pulls Grand Totals from each department tab. Do not edit directly."
c.font = Font(name="Arial", italic=True, color="555555", size=9)
c.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
c.alignment = Alignment(horizontal="center", vertical="center")

ws_sum.row_dimensions[3].height = 26
hdr(ws_sum["A3"], "#")
hdr(ws_sum["B3"], "Section / Period", align="left")
for i, dept in enumerate(DEPARTMENTS, start=3):
    # Truncated dept name for column header
    short = dept[:20]
    hdr(ws_sum.cell(row=3, column=i), short, bg=MID_BLUE, size=8, wrap=True)
hdr(ws_sum.cell(row=3, column=end_col), "GRAND TOTAL", bg=HEADER_GOLD, fg=DARK_NAVY)

dept_grand_total_rows = {}

# Build dept sheets
for dept in DEPARTMENTS:
    sec_totals, gt_row = build_dept(wb, dept)
    dept_grand_total_rows[dept] = (gt_row, sec_totals)

# Default state before macros run:
# keep only instructions visible; hide summary + all department sheets.
wb["CONSOLIDATED SUMMARY"].sheet_state = "veryHidden"
for dept in DEPARTMENTS:
    wb[dept[:31]].sheet_state = "veryHidden"

# Fill summary rows — detailed monthly + FY per section
row = 4
period_labels = MONTHS + ["FY Total"]
for sec_idx, (sec_title, _) in enumerate(SECTIONS, start=1):
    sec_label = sec_title.split(". ", 1)[1] if ". " in sec_title else sec_title

    # Section header row
    ws_sum.row_dimensions[row].height = 20
    ws_sum.cell(row=row, column=1).value = str(sec_idx)
    ws_sum.cell(row=row, column=1).fill = PatternFill("solid", fgColor=SECTION_BG)
    ws_sum.cell(row=row, column=1).border = thin()
    ws_sum.cell(row=row, column=1).font = Font(name="Arial", bold=True, size=9, color="333333")
    lbl(ws_sum.cell(row=row, column=2), sec_label, bold=True, bg=SECTION_BG, indent=1)

    for d_idx, dept in enumerate(DEPARTMENTS):
        col = 3 + d_idx
        c = ws_sum.cell(row=row, column=col)
        c.fill = PatternFill("solid", fgColor=SECTION_BG)
        c.border = thin()

    ws_sum.cell(row=row, column=end_col).fill = PatternFill("solid", fgColor=SECTION_BG)
    ws_sum.cell(row=row, column=end_col).border = thin()
    row += 1

    # One row per month + FY total
    for p_idx, period in enumerate(period_labels):
        ws_sum.row_dimensions[row].height = 18
        ws_sum.cell(row=row, column=1).value = ""
        ws_sum.cell(row=row, column=1).fill = PatternFill("solid", fgColor=ALT_ROW if row % 2 == 0 else WHITE)
        ws_sum.cell(row=row, column=1).border = thin()
        lbl(ws_sum.cell(row=row, column=2), period, bold=(period == "FY Total"),
            bg=ALT_ROW if row % 2 == 0 else WHITE, indent=2)

        for d_idx, dept in enumerate(DEPARTMENTS):
            _, sec_totals = dept_grand_total_rows[dept]
            sname = dept[:31]
            matched = None
            for sk, sr in sec_totals.items():
                if sec_label.lower() in sk.lower():
                    matched = sr
                    break

            col = 3 + d_idx
            c = ws_sum.cell(row=row, column=col)
            if matched:
                if period == "FY Total":
                    formula = f"='{sname}'!{get_column_letter(16)}{matched}"
                else:
                    month_col = 4 + p_idx
                    formula = f"='{sname}'!{get_column_letter(month_col)}{matched}"
                frm(c, formula)
            else:
                c.value = 0
                c.fill = PatternFill("solid", fgColor=FORMULA_BG)
                c.border = thin()
                c.number_format = '#,##0;(#,##0);"-"'

        # Row total across departments
        start_c = get_column_letter(3)
        end_c = get_column_letter(2 + len(DEPARTMENTS))
        if period == "FY Total":
            tot(ws_sum.cell(row=row, column=end_col), f"=SUM({start_c}{row}:{end_c}{row})", bg=ACCENT_TEAL)
            ws_sum.cell(row=row, column=end_col).font = Font(name="Arial", bold=True, color=WHITE, size=9)
            ws_sum.cell(row=row, column=end_col).fill = PatternFill("solid", fgColor=ACCENT_TEAL)
            ws_sum.cell(row=row, column=2).fill = PatternFill("solid", fgColor=TOTAL_BG)
        else:
            tot(ws_sum.cell(row=row, column=end_col), f"=SUM({start_c}{row}:{end_c}{row})", bg=FORMULA_BG, bold=False)

        row += 1

    blank(ws_sum, row, 1, end_col)
    row += 1

# Grand Total row in summary
ws_sum.row_dimensions[row].height = 28
ws_sum.merge_cells(f"A{row}:B{row}")
c = ws_sum[f"A{row}"]
c.value = "  TOTAL — ALL DEPARTMENTS"
c.font = Font(name="Arial", bold=True, color=WHITE, size=11)
c.fill = PatternFill("solid", fgColor=DARK_NAVY)
c.alignment = Alignment(horizontal="left", vertical="center")
c.border = thick()

for d_idx, dept in enumerate(DEPARTMENTS):
    gt_row, _ = dept_grand_total_rows[dept]
    sname = dept[:31]
    col = 3 + d_idx
    c2 = ws_sum.cell(row=row, column=col)
    c2.value = (f"=SUM('{sname}'!{get_column_letter(4)}{gt_row}:"
                f"'{sname}'!{get_column_letter(15)}{gt_row})")
    c2.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c2.fill = PatternFill("solid", fgColor=DARK_NAVY)
    c2.border = thick()
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.number_format = '#,##0;(#,##0);"-"'

gt_final = ws_sum.cell(row=row, column=end_col)
gt_final.value = f"=SUM({get_column_letter(3)}{row}:{get_column_letter(2+len(DEPARTMENTS))}{row})"
gt_final.font = Font(name="Arial", bold=True, color=WHITE, size=11)
gt_final.fill = PatternFill("solid", fgColor=DARK_NAVY)
gt_final.border = thick()
gt_final.alignment = Alignment(horizontal="center", vertical="center")
gt_final.number_format = '#,##0;(#,##0);"-"'

ws_sum.freeze_panes = "C4"

# ─── SAVE AS .xlsx ────────────────────────────────────────────────────────────
# openpyxl cannot embed compiled VBA into a new workbook.
# We generate a valid .xlsx and a companion .bas file for one-time import.

os.makedirs(OUTPUT_DIR, exist_ok=True)
xlsx_out = os.path.join(OUTPUT_DIR, OUTPUT_FILE)
generated_workbook_name = os.path.basename(xlsx_out)

try:
    wb.save(xlsx_out)
except PermissionError:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    generated_workbook_name = f"Budget_FY2027_MGA_{stamp}.xlsx"
    xlsx_out = os.path.join(OUTPUT_DIR, generated_workbook_name)
    wb.save(xlsx_out)

# ─── BUILD THE VBA MODULE ────────────────────────────────────────────────────
# Build password lookup string for VBA
pw_lines = []
for dept, pw in DEPT_PASSWORDS.items():
    sname = dept[:31].replace('"', '""')
    pw_lines.append(f'    Case "{pw}": targetSheet = "{sname}"')

pw_lookup = "\n".join(pw_lines)

dept_sheets_list = '", "'.join([d[:31] for d in DEPARTMENTS])

VBA_CODE = f'''
Attribute VB_Name = "BudgetAccess"
Option Explicit

Public Sub Auto_Open()
    Call BudgetAccess_Open
End Sub

Public Sub BudgetAccess_Open()
    Dim enteredPW As String
    Dim targetSheet As String
    Dim ws As Worksheet
    Dim masterPW As String
    
    masterPW = "{MASTER_PASSWORD}"
    
    ' Hide ALL department sheets first
    Call HideAllDeptSheets
    
    ' Prompt for password
    enteredPW = InputBox( _
        "Enter your department budget password:" & Chr(13) & Chr(10) & Chr(13) & Chr(10) & _
        "Contact Finance if you do not have your password.", _
        "MGA Apparel — Budget FY 2026-27", "")
    
    If enteredPW = "" Then
        MsgBox "No password entered. Department and summary sheets remain hidden.", vbInformation, "Budget Access"
        Exit Sub
    End If
    
    ' Check master password
    If enteredPW = masterPW Then
        Call ShowAllDeptSheets
        MsgBox "Master access granted. All department sheets are now visible.", vbInformation, "Master Access"
        Exit Sub
    End If
    
    ' Check department passwords
    Select Case enteredPW
{pw_lookup}
    Case Else
        targetSheet = ""
    End Select
    
    If targetSheet = "" Then
        MsgBox "Incorrect password. Please contact Finance for your department password.", _
               vbExclamation, "Access Denied"
        Exit Sub
    End If
    
    ' Show only the matching dept sheet
    For Each ws In ThisWorkbook.Sheets
        If ws.Name = targetSheet Then
            ws.Visible = xlSheetVisible
            ws.Activate
        End If
    Next ws
    
    MsgBox "Welcome! Your department sheet """ & targetSheet & """ is now unlocked.", _
           vbInformation, "Budget FY 2026-27"
End Sub

Private Sub HideAllDeptSheets()
    Dim ws As Worksheet
    Dim deptSheets As Variant
    deptSheets = Array("{dept_sheets_list}")
    On Error Resume Next
    ThisWorkbook.Sheets("CONSOLIDATED SUMMARY").Visible = xlSheetVeryHidden
    On Error GoTo 0
    Dim i As Integer
    For Each ws In ThisWorkbook.Sheets
        Dim isDept As Boolean
        isDept = False
        For i = 0 To UBound(deptSheets)
            If ws.Name = deptSheets(i) Then
                isDept = True
                Exit For
            End If
        Next i
        If isDept Then
            ws.Visible = xlSheetVeryHidden
        End If
    Next ws
End Sub

Private Sub ShowAllDeptSheets()
    Dim ws As Worksheet
    Dim deptSheets As Variant
    deptSheets = Array("{dept_sheets_list}")
    On Error Resume Next
    ThisWorkbook.Sheets("CONSOLIDATED SUMMARY").Visible = xlSheetVisible
    On Error GoTo 0
    Dim i As Integer
    For i = 0 To UBound(deptSheets)
        On Error Resume Next
        ThisWorkbook.Sheets(deptSheets(i)).Visible = xlSheetVisible
        On Error GoTo 0
    Next i
    ' Navigate to summary
    ThisWorkbook.Sheets("CONSOLIDATED SUMMARY").Activate
End Sub
'''

# ─── VBA DELIVERY ─────────────────────────────────────────────────────────────
# Full VBA embedding requires an existing vbaProject.bin template.
# The cross-platform approach is to generate a .bas file for one-time import,
# then save the workbook as .xlsm inside Excel.

# Save a companion .bas file the user imports once
bas_path = os.path.join(OUTPUT_DIR, "ImportThisVBA_OneTime.bas")
with open(bas_path, "w", encoding="utf-8") as f:
    f.write(VBA_CODE)

# ─── WRITE PASSWORD REFERENCE CSV ────────────────────────────────────────────
csv_path = os.path.join(OUTPUT_DIR, "DepartmentPasswords_CONFIDENTIAL.csv")
with open(csv_path, "w", encoding="utf-8") as f:
    f.write("Department,Password\n")
    f.write(f"[MASTER],{MASTER_PASSWORD}\n")
    for dept, pw in DEPT_PASSWORDS.items():
        f.write(f'"{dept}",{pw}\n')

# ─── WRITE SETUP INSTRUCTIONS ────────────────────────────────────────────────
readme_path = os.path.join(OUTPUT_DIR, "SETUP_INSTRUCTIONS.txt")
with open(readme_path, "w", encoding="utf-8") as f:
    f.write("""
=============================================================
  MGA BUDGET FY2027 — SETUP INSTRUCTIONS (Finance Team Only)
=============================================================

Files generated in this folder:
    - """ + generated_workbook_name + """         ← Base budget file (convert to .xlsm after VBA import)
  - ImportThisVBA_OneTime.bas      ← VBA macro to import ONCE (see below)
  - DepartmentPasswords_CONFIDENTIAL.csv  ← Keep this secure!
  - SETUP_INSTRUCTIONS.txt         ← This file

─── ONE-TIME VBA SETUP (do this before sharing) ──────────────
You need to import the VBA macro into the Excel file once.
openpyxl cannot embed compiled VBA, so this is a manual step:

1. Open """ + generated_workbook_name + """ in Excel (desktop app, not browser)
2. Press Alt + F11 to open the VBA Editor
3. In the Project pane (left), right-click on the workbook project for this file
4. Click "Import File..."
5. Browse to ImportThisVBA_OneTime.bas and click Open
6. Close the VBA Editor (Alt+F11 again)
7. Save the file as .xlsm (File > Save As > Excel Macro-Enabled Workbook)
8. Test: Close and reopen the file. A password box should appear.
9. If no prompt appears, press Alt+F8 and run: Auto_Open

─── MACRO SECURITY NOTE ──────────────────────────────────────
When sharing via OneDrive:
- Recipients must click "Enable Macros" when opening
- If they open in Excel Online (browser), macros won't run
  → They MUST open in the desktop Excel app
- You can add a note in the email: "Please open in Excel Desktop App"

─── SHARING ──────────────────────────────────────────────────
1. Upload Budget_FY2027_MGA.xlsm to your OneDrive folder
2. Share the OneDrive LINK with each department (not the file itself)
3. Send each department their password separately (via email/WhatsApp)
4. Keep DepartmentPasswords_CONFIDENTIAL.csv in a secure location only

─── MASTER PASSWORD ──────────────────────────────────────────
  Master Password: """ + MASTER_PASSWORD + """
  This unlocks ALL department sheets simultaneously.
  Use it to review all data and check the Consolidated Summary.

=============================================================
""")

print("=" * 60)
print("Budget workbook generated successfully!")
print(f"Output directory: {OUTPUT_DIR}")
print()
print("Files created:")
print(f"  {generated_workbook_name}  (import VBA, then Save As .xlsm before sharing)")
print(f"  ImportThisVBA_OneTime.bas  (import into Excel once)")
print(f"  DepartmentPasswords_CONFIDENTIAL.csv")
print(f"  SETUP_INSTRUCTIONS.txt")
print()
print("NEXT STEP: Read SETUP_INSTRUCTIONS.txt before sharing!")
print("=" * 60)