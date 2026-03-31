import json
import sqlite3
import uuid
import base64
import os
from datetime import datetime
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

import pandas as pd
import requests
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

APP_TITLE = "MGA Budget Portal FY2026-27"
DB_PATH = Path(__file__).with_name("budget_data.db")
PASSWORD_CSV = Path(__file__).with_name("DepartmentPasswords_CONFIDENTIAL.csv")
APP_SETTINGS_KEY = "__APP_SETTINGS__"
LOGO_CANDIDATES = [
    "mg_apparel_logo.png",
    "mg_apparel_logo.jpg",
    "mg_apparel_logo.jpeg",
    "MG Apparel Logo.png",
    "MG Apparel Logo.jpg",
    "MG Apparel Logo.jpeg",
    "logo.png",
    "logo.jpg",
    "logo.jpeg",
]
SUPABASE_URL = os.getenv("SUPABASE_URL", "").strip().rstrip("/")
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "").strip()
USE_SUPABASE = bool(SUPABASE_URL and SUPABASE_KEY)
MONTHS = [
    "Jul-26", "Aug-26", "Sep-26", "Oct-26", "Nov-26", "Dec-26",
    "Jan-27", "Feb-27", "Mar-27", "Apr-27", "May-27", "Jun-27",
]
SECTION_TEMPLATES = [
    ("1. TRAINING & DEVELOPMENT", [("Employee 1", "units_rate"), ("Employee 2", "units_rate"), ("Employee 3", "units_rate")]),
    ("2. SOFTWARE & LICENSES", [("Platform A", "units_rate"), ("Platform B", "units_rate"), ("Platform C", "units_rate")]),
    ("3. IT EQUIPMENT (CAPEX)", [("Laptops", "units_rate"), ("Others", "units_rate")]),
    ("4. STORE CONSUMPTION", [("Pens", "units_rate"), ("Paper (Ream)", "units_rate"), ("Toner/Ink", "units_rate"), ("Other Items", "units_rate")]),
    ("5. ENTERTAINMENT", [("Client Entertainment", "units_rate"), ("Staff Events", "units_rate")]),
    ("6. FOREIGN TRAVEL (USD)", [("Airfare", "units_rate"), ("Food & Lodging", "units_rate"), ("Other Costs", "amount")]),
    ("7. LOCAL TRAVEL (PKR)", [("Multan", "units_rate"), ("Lahore", "units_rate"), ("Islamabad", "units_rate"), ("Karachi", "units_rate"), ("Other", "units_rate")]),
]
SECTION_NAMES = [s for s, _ in SECTION_TEMPLATES]
MAX_ATTACHMENT_BYTES = 5 * 1024 * 1024
ALLOWED_ATTACHMENT_TYPES = {
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "image/png",
    "image/jpeg",
}

EMBED_SHEETS_SECRET_KEY = "EMBED_SHEETS"

TRAVEL_COST_KEYS = ["Food", "Lodging", "Travelling Fare"]

SUPPLY_CHAIN_DOMAIN = "SUPPLY CHAIN"
SUPPLY_CHAIN_SEGMENTS = [
    "FABRIC SOURCING",
    "EXPORT & LOGISTICS",
    "MATERIAL MANAGEMENT & CONTROL",
]

MARKETING_DEPT = "MARKETING & MERCHANDIZING"
PPC_DEPT = "PPC & WIP"

_HERE = Path(__file__).resolve().parent
MKT_TEMPLATES_DIR = (_HERE / "MKT") if (_HERE / "MKT").exists() else (_HERE.parent / "MKT")
MKT_TEMPLATE_FILES = {
    "Sales Plan.xlsx": {
        "title": "Sales Plan",
        "shared": False,
    },
    "Working Capital.xlsx": {
        "title": "Working Capital",
        "shared": True,
    },
}

REMOVED_DEPARTMENTS = {
    "R61 OPERATIONS",
}

MASTER_DOMAIN_NAME = "Hamza Zahid"
DOMAINS_XLSX = Path(__file__).with_name("Domains.xlsx")
DOMAINS_CSV = Path(__file__).with_name("Domains.csv")
DOMAINS_DIR = Path(__file__).with_name("Domains")

DEFAULT_DEPARTMENT_DOMAINS = {
    "ACCOUNTS": "Umer Malik",
    "AUDIT": "Usman BK",
    "BUSINESS AFFAIRS, SUSTAINABILITY, CSR": "Kehkeshan",
    "PD & SAMPLING": "Alaoudin",
    "MARKETING & MERCHANDIZING": "Naqeeb",
    "RESEARCH & DESIGN": "Naqeeb",
    "SUPPLY CHAIN": "Kashif Basit",
    "ENGINEERING & UTILITIES": "Alaoudin",
    "ADMINISTRATION": "Saman",
    "COMPLIANCE, HSE & IR": "Saman",
    "STORES": "Kashif Basit",
    "HUMAN RESOURCES": "Saman",
    "CUTTING & EMBROIDERY": "Alaoudin",
    "STITCHING": "Alaoudin",
    "WASHING & DRY PROCESS": "Alaoudin",
    "FINISHING": "Alaoudin",
    "PPC & WIP": "Alaoudin",
    "IE & PROCESS IMPROVEMENT": "Alaoudin",
    "MAINTENANCE": "Alaoudin",
    "MIS & IT": "Ahmed",
    "QUALITY ASSURANCE": "Alaoudin",
    "QUALITY CONTROL": "Alaoudin",
    "DSBA": "Hamza Zahid",
}


def _normalize_domain(value: str) -> str:
    return " ".join(str(value or "").strip().split())


def _load_domains_from_table(df: pd.DataFrame) -> dict[str, str]:
    if df is None or df.empty:
        return {}

    cols = {c: str(c).strip().lower() for c in df.columns}
    dept_col = next((c for c, v in cols.items() if v in {"dept", "department"}), None)
    dom_col = next((c for c, v in cols.items() if v in {"domain"}), None)
    if not dept_col or not dom_col:
        return {}

    mapping: dict[str, str] = {}
    for _, row in df.iterrows():
        dept = normalize_name(str(row.get(dept_col, "")))
        dom = _normalize_domain(str(row.get(dom_col, "")))
        if not dept or not dom:
            continue
        if dept in REMOVED_DEPARTMENTS:
            continue
        mapping[dept] = dom
    return mapping


def _load_domains_from_filesystem() -> dict[str, str]:
    # Preferred: a file shipped with the app repo.
    paths: list[Path] = []
    if DOMAINS_DIR.exists() and DOMAINS_DIR.is_dir():
        for p in sorted(DOMAINS_DIR.iterdir()):
            if p.suffix.lower() in {".xlsx", ".csv"}:
                paths.append(p)
    for p in [DOMAINS_XLSX, DOMAINS_CSV]:
        if p.exists():
            paths.append(p)

    for p in paths:
        try:
            if p.suffix.lower() == ".csv":
                df = pd.read_csv(p)
            elif p.suffix.lower() == ".xlsx":
                df = pd.read_excel(p)
            else:
                continue
        except Exception:
            continue

        mapping = _load_domains_from_table(df)
        if mapping:
            return mapping

    return {}


def load_department_domains() -> dict[str, str]:
    """Returns a mapping of DEPARTMENT -> DOMAIN.

    Load order:
    1) Repo file(s): Domains.xlsx / Domains.csv / Domains/*
    2) Streamlit Secrets: DEPARTMENT_DOMAINS
    3) Built-in defaults
    """
    mapping = _load_domains_from_filesystem()

    if not mapping:
        try:
            raw = st.secrets.get("DEPARTMENT_DOMAINS", {})
            raw_items = dict(raw).items()
        except Exception:
            raw_items = []

        for dept, dom in raw_items:
            d = normalize_name(str(dept))
            v = _normalize_domain(str(dom))
            if d and v and d not in REMOVED_DEPARTMENTS:
                mapping[d] = v

    if not mapping:
        for dept, dom in DEFAULT_DEPARTMENT_DOMAINS.items():
            d = normalize_name(dept)
            v = _normalize_domain(dom)
            if d and v and d not in REMOVED_DEPARTMENTS:
                mapping[d] = v

    # Derive Supply Chain domain from the segments if the file doesn't include SUPPLY CHAIN row.
    if SUPPLY_CHAIN_DOMAIN not in mapping:
        seg_domains = {_normalize_domain(mapping.get(seg, "")) for seg in SUPPLY_CHAIN_SEGMENTS}
        seg_domains = {d for d in seg_domains if d}
        if len(seg_domains) == 1:
            mapping[SUPPLY_CHAIN_DOMAIN] = next(iter(seg_domains))

    # Ensure segments inherit Supply Chain domain if available.
    sc_dom = mapping.get(SUPPLY_CHAIN_DOMAIN)
    if sc_dom:
        for seg in SUPPLY_CHAIN_SEGMENTS:
            mapping.setdefault(seg, sc_dom)

    return mapping


def department_domain_for_login(dept_domains: dict[str, str], login_as: str) -> str:
    if login_as == SUPPLY_CHAIN_DOMAIN:
        dom = dept_domains.get(SUPPLY_CHAIN_DOMAIN, "").strip()
        if dom:
            return dom
        seg_domains = {str(dept_domains.get(seg, "")).strip() for seg in SUPPLY_CHAIN_SEGMENTS}
        seg_domains = {d for d in seg_domains if d}
        if len(seg_domains) == 1:
            return next(iter(seg_domains))
        return ""
    return str(dept_domains.get(login_as, "")).strip()


def supply_chain_enabled(auth_map: dict) -> bool:
    return SUPPLY_CHAIN_DOMAIN in auth_map


def login_domains(auth_map: dict) -> list[str]:
    domains = sorted(auth_map.keys())
    if supply_chain_enabled(auth_map):
        domains = [d for d in domains if d not in SUPPLY_CHAIN_SEGMENTS and d != SUPPLY_CHAIN_DOMAIN]
        domains.append(SUPPLY_CHAIN_DOMAIN)
        domains = sorted(domains)
    return domains


def check_domain_password(auth_map: dict, domain: str, password: str) -> bool:
    if domain == SUPPLY_CHAIN_DOMAIN and supply_chain_enabled(auth_map):
        return auth_map.get(SUPPLY_CHAIN_DOMAIN) == password
    return auth_map.get(domain) == password


def default_app_settings() -> dict:
    return {
        "edit_locked": False,
        "view_locked": False,
        "locked_by": None,
        "locked_at": None,
        "updated_at": None,
    }


def normalize_name(name: str) -> str:
    return name.strip().upper()


def _get_logo_source() -> str | Path | None:
    try:
        logo_url = str(st.secrets.get("LOGO_URL", "")).strip()
    except Exception:
        logo_url = ""
    if logo_url:
        return logo_url

    for name in LOGO_CANDIDATES:
        candidate = Path(__file__).with_name(name)
        if candidate.exists():
            return candidate
    return None


def render_header(compact: bool = False) -> None:
    logo = _get_logo_source()
    if logo:
        c1, c2, c3 = st.columns([2, 1, 2])
        with c2:
            st.image(logo, use_container_width=True)

    if compact:
        st.markdown(f"## {APP_TITLE}")
    else:
        st.markdown(f"# {APP_TITLE}")
        st.caption("Login with your assigned department password.")


def month_map(default: float = 0.0) -> dict:
    return {m: float(default) for m in MONTHS}


def travel_breakdown_default() -> dict:
    return {k: month_map(0.0) for k in TRAVEL_COST_KEYS}


def ensure_travel_breakdown(item: dict) -> None:
    tb = item.get("travel_breakdown")
    if not isinstance(tb, dict):
        tb = travel_breakdown_default()
        item["travel_breakdown"] = tb

    for k in TRAVEL_COST_KEYS:
        tb.setdefault(k, month_map(0.0))
        for m in MONTHS:
            tb[k][m] = _to_float(tb[k].get(m, 0.0))


def create_item(name: str, kind: str = "units_rate", item_id: str | None = None) -> dict:
    return {
        "id": item_id or uuid.uuid4().hex,
        "name": name,
        "description": "",
        "kind": kind,
        "units": month_map(0.0),
        "rate": month_map(0.0),
        "amount": month_map(0.0),
        "travel_breakdown": None,
        "benefit": month_map(0.0),
        "comment": "",
        "attachments": [],
    }


def attachment_label(att: dict) -> str:
    size_kb = int((_to_float(att.get("size", 0)) + 1023) // 1024)
    return f"{att.get('name', 'document')} ({size_kb} KB)"


def get_embed_sheet_url(doc_file: str) -> str:
    """Fetch an Excel Online embed URL from Streamlit Secrets.

    Expected secrets shape:
      [EMBED_SHEETS]
      "Working Capital.xlsx" = "https://.../embed?..."
      "Sales Plan.xlsx" = "https://.../embed?..."
    """
    try:
        raw = st.secrets.get(EMBED_SHEETS_SECRET_KEY, {})
        items = dict(raw).items()
    except Exception:
        items = []

    wanted = doc_file.strip().lower()
    for k, v in items:
        if str(k).strip().lower() == wanted:
            return str(v).strip()
    return ""


def can_access_shared_sheet(role: str, dept: str | None, dept_domain: str | None, doc_file: str) -> bool:
    if role == "master":
        return True

    if not dept:
        return False

    if doc_file not in MKT_TEMPLATE_FILES:
        return False

    cfg = MKT_TEMPLATE_FILES[doc_file]
    if not cfg.get("shared", False):
        return dept == MARKETING_DEPT

    # Working Capital: Marketing, PPC & WIP, Supply Chain domain.
    if dept == MARKETING_DEPT or dept == PPC_DEPT:
        return True
    if (dept_domain or "").strip().upper() == SUPPLY_CHAIN_DOMAIN:
        return True
    return False


def render_shared_sheets_panel(edit_locked: bool, view_locked: bool) -> None:
    role = st.session_state.role
    dept = st.session_state.department
    dept_domain = st.session_state.get("department_domain")
    if role != "master" and not dept:
        return
    if role != "master" and view_locked:
        return

    visible_files = [
        f for f in MKT_TEMPLATE_FILES.keys()
        if can_access_shared_sheet(role, dept, dept_domain, f)
    ]
    if not visible_files:
        return

    st.subheader("Sheets")
    st.caption("These sheets open inside the portal in Excel Online (formats & formulas unchanged).")

    for doc_file in visible_files:
        cfg = MKT_TEMPLATE_FILES[doc_file]
        title = cfg.get("title", doc_file)
        is_shared = bool(cfg.get("shared", False))
        url = get_embed_sheet_url(doc_file)

        with st.expander(f"{title}{' (Shared)' if is_shared else ''}", expanded=False):
            if not url:
                st.error("Sheet link not configured. Add an Excel embed URL in Streamlit Secrets under [EMBED_SHEETS].")
                st.code(
                    """[EMBED_SHEETS]\n\"Working Capital.xlsx\" = \"https://.../embed?...\"\n\"Sales Plan.xlsx\" = \"https://.../embed?...\"\n""",
                    language="toml",
                )
                continue

            if edit_locked and role != "master":
                st.info("Note: Department editing is locked in the portal. Sheet editing permissions are controlled by OneDrive/SharePoint.")

            st.markdown(f"Open in new tab: {url}")
            components.iframe(url, height=720, scrolling=True)


def _is_duplicate_attachment(item: dict, uploaded_name: str, uploaded_size: int) -> bool:
    for att in item.get("attachments", []):
        if att.get("name") == uploaded_name and int(_to_float(att.get("size", 0))) == int(uploaded_size):
            return True
    return False


def add_attachment(item: dict, uploaded_file) -> str | None:
    if uploaded_file is None:
        return None

    file_bytes = uploaded_file.getvalue()
    if len(file_bytes) > MAX_ATTACHMENT_BYTES:
        return "File too large. Max allowed size is 5 MB."
    if uploaded_file.type not in ALLOWED_ATTACHMENT_TYPES:
        return "Unsupported file type. Use PDF, XLSX, DOCX, PNG, or JPG."
    if _is_duplicate_attachment(item, uploaded_file.name, len(file_bytes)):
        return None

    item.setdefault("attachments", []).append(
        {
            "id": uuid.uuid4().hex,
            "name": uploaded_file.name,
            "mime": uploaded_file.type,
            "size": len(file_bytes),
            "content_b64": base64.b64encode(file_bytes).decode("ascii"),
            "uploaded_at": datetime.now().isoformat(timespec="seconds"),
        }
    )
    return None


def default_payload() -> dict:
    payload = {"sections": {}, "updated_at": None}
    for section, items in SECTION_TEMPLATES:
        payload["sections"][section] = [create_item(item_name, kind) for item_name, kind in items]
    return payload


def ensure_db() -> None:
    if USE_SUPABASE:
        return

    conn = sqlite3.connect(DB_PATH)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS budget_entries (
            department TEXT PRIMARY KEY,
            payload_json TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.commit()
    conn.close()


def _supabase_headers() -> dict:
    return {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }


def _supabase_get_payload(department: str):
    dept_encoded = quote(department, safe="")
    url = (
        f"{SUPABASE_URL}/rest/v1/budget_entries"
        f"?select=payload_json&department=eq.{dept_encoded}&limit=1"
    )
    resp = requests.get(url, headers=_supabase_headers(), timeout=20)
    if resp.status_code >= 400:
        raise RuntimeError(f"Supabase read failed: {resp.status_code} {resp.text[:200]}")
    data = resp.json()
    if not data:
        return None
    return data[0].get("payload_json")


def _supabase_upsert_payload(department: str, payload: dict, updated_at: str) -> None:
    url = f"{SUPABASE_URL}/rest/v1/budget_entries"
    body = [{"department": department, "payload_json": payload, "updated_at": updated_at}]
    headers = _supabase_headers()
    headers["Prefer"] = "resolution=merge-duplicates,return=minimal"
    resp = requests.post(url, headers=headers, data=json.dumps(body), timeout=20)
    if resp.status_code >= 400:
        raise RuntimeError(f"Supabase write failed: {resp.status_code} {resp.text[:200]}")


def _load_auth_from_secrets() -> tuple[dict, str] | None:
    try:
        secrets = st.secrets
    except Exception:
        return None

    master_password = str(secrets.get("MASTER_PASSWORD", "")).strip()
    dept_block = secrets.get("DEPARTMENT_PASSWORDS", {})
    try:
        dept_items = dict(dept_block).items()
    except Exception:
        dept_items = []

    dept_passwords = {}
    for dept, pw in dept_items:
        d = normalize_name(str(dept))
        p = str(pw).strip()
        if d and p:
            dept_passwords[d] = p

    if master_password and dept_passwords:
        return dept_passwords, master_password
    return None


def load_auth_map() -> tuple[dict, str]:
    from_secrets = _load_auth_from_secrets()
    if from_secrets:
        dept_passwords, master_pw = from_secrets
        for d in REMOVED_DEPARTMENTS:
            dept_passwords.pop(normalize_name(d), None)
        return dept_passwords, master_pw

    if not PASSWORD_CSV.exists():
        st.error("No auth source found. Configure Streamlit Secrets or provide DepartmentPasswords_CONFIDENTIAL.csv")
        st.stop()

    df = pd.read_csv(PASSWORD_CSV)
    dept_passwords = {}
    master_password = ""

    for _, row in df.iterrows():
        dept = str(row["Department"]).strip()
        pw = str(row["Password"]).strip()
        if dept == "[MASTER]":
            master_password = pw
        else:
            dept_key = normalize_name(dept)
            dept_passwords[dept_key] = pw

    if not master_password:
        st.error("Master password entry [MASTER] missing in DepartmentPasswords_CONFIDENTIAL.csv")
        st.stop()

    for d in REMOVED_DEPARTMENTS:
        dept_passwords.pop(normalize_name(d), None)

    return dept_passwords, master_password


def _to_float(value) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def migrate_payload(payload: dict) -> dict:
    if payload.get("sections"):
        for section in SECTION_NAMES:
            payload["sections"].setdefault(section, [])
        for section, items in payload["sections"].items():
            for item in items:
                item.setdefault("id", uuid.uuid4().hex)
                item.setdefault("name", "New Subhead")
                item.setdefault("description", "")
                item.setdefault("kind", "units_rate")
                item.setdefault("units", month_map(0.0))
                item.setdefault("rate", month_map(0.0))
                item.setdefault("amount", month_map(0.0))
                item.setdefault("travel_breakdown", None)
                item.setdefault("benefit", month_map(0.0))
                item.setdefault("comment", "")
                item.setdefault("attachments", [])

                if is_travel_section(section):
                    if not isinstance(item.get("travel_breakdown"), dict):
                        existing_costs = []
                        if item.get("kind") == "amount":
                            existing_costs = [_to_float(item["amount"].get(m, 0.0)) for m in MONTHS]
                        else:
                            units = [_to_float(item["units"].get(m, 0.0)) for m in MONTHS]
                            rate = [_to_float(item["rate"].get(m, 0.0)) for m in MONTHS]
                            existing_costs = [units[i] * rate[i] for i in range(len(MONTHS))]

                        item["travel_breakdown"] = travel_breakdown_default()
                        for i, m in enumerate(MONTHS):
                            item["travel_breakdown"]["Travelling Fare"][m] = _to_float(existing_costs[i] if i < len(existing_costs) else 0.0)

                    ensure_travel_breakdown(item)
                    item["kind"] = "travel_breakdown"

                for m in MONTHS:
                    item["units"][m] = _to_float(item["units"].get(m, 0.0))
                    item["rate"][m] = _to_float(item["rate"].get(m, 0.0))
                    item["amount"][m] = _to_float(item["amount"].get(m, 0.0))
                    item["benefit"][m] = _to_float(item["benefit"].get(m, 0.0))
        payload.setdefault("updated_at", None)
        return payload

    old_rows = payload.get("rows", {})
    old_comments = payload.get("comments", {})
    new_payload = default_payload()

    for section, _ in SECTION_TEMPLATES:
        section_items = new_payload["sections"][section]
        for item in section_items:
            name = item["name"]
            units_key = f"{section}||{name}||units"
            rate_key = f"{section}||{name}||rate"
            amount_key = f"{section}||{name}||other"
            comment_key = f"{section}||{name}||comment"

            if amount_key in old_rows:
                item["kind"] = "amount"
                for m in MONTHS:
                    item["amount"][m] = _to_float(old_rows.get(amount_key, {}).get(m, 0.0))
            else:
                item["kind"] = "units_rate"
                for m in MONTHS:
                    item["units"][m] = _to_float(old_rows.get(units_key, {}).get(m, 0.0))
                    item["rate"][m] = _to_float(old_rows.get(rate_key, {}).get(m, 0.0))

            if is_travel_section(section):
                existing_costs = item_cost_by_month(item)
                item["travel_breakdown"] = travel_breakdown_default()
                for i, m in enumerate(MONTHS):
                    item["travel_breakdown"]["Travelling Fare"][m] = _to_float(existing_costs[i] if i < len(existing_costs) else 0.0)
                item["kind"] = "travel_breakdown"

            item["comment"] = str(old_comments.get(comment_key, ""))

    new_payload["updated_at"] = payload.get("updated_at")
    return new_payload


def load_app_settings() -> dict:
    defaults = default_app_settings()

    if USE_SUPABASE:
        raw = _supabase_get_payload(APP_SETTINGS_KEY)
        if not raw:
            return defaults
        if isinstance(raw, str):
            try:
                raw = json.loads(raw)
            except Exception:
                return defaults
        if not isinstance(raw, dict):
            return defaults
        merged = {**defaults, **raw}
        merged["edit_locked"] = bool(merged.get("edit_locked", False))
        merged["view_locked"] = bool(merged.get("view_locked", False))
        return merged

    conn = sqlite3.connect(DB_PATH)
    row = conn.execute(
        "SELECT payload_json FROM budget_entries WHERE department = ?", (APP_SETTINGS_KEY,)
    ).fetchone()
    conn.close()

    if not row:
        return defaults

    try:
        raw = json.loads(row[0])
    except Exception:
        return defaults
    if not isinstance(raw, dict):
        return defaults
    merged = {**defaults, **raw}
    merged["edit_locked"] = bool(merged.get("edit_locked", False))
    merged["view_locked"] = bool(merged.get("view_locked", False))
    return merged


def save_app_settings(settings: dict) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    payload = {**default_app_settings(), **(settings or {})}
    payload["updated_at"] = now

    if USE_SUPABASE:
        _supabase_upsert_payload(APP_SETTINGS_KEY, payload, now)
        return

    conn = sqlite3.connect(DB_PATH)
    conn.execute(
        """
        INSERT INTO budget_entries (department, payload_json, updated_at)
        VALUES (?, ?, ?)
        ON CONFLICT(department)
        DO UPDATE SET payload_json=excluded.payload_json, updated_at=excluded.updated_at
        """,
        (APP_SETTINGS_KEY, json.dumps(payload), now),
    )
    conn.commit()
    conn.close()


def save_payload(department: str, payload: dict) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    payload["updated_at"] = now

    if USE_SUPABASE:
        _supabase_upsert_payload(department, payload, now)
        return

    conn = sqlite3.connect(DB_PATH)
    conn.execute(
        """
        INSERT INTO budget_entries (department, payload_json, updated_at)
        VALUES (?, ?, ?)
        ON CONFLICT(department)
        DO UPDATE SET payload_json=excluded.payload_json, updated_at=excluded.updated_at
        """,
        (department, json.dumps(payload), now),
    )
    conn.commit()
    conn.close()


def load_payload(department: str) -> dict:
    if USE_SUPABASE:
        raw = _supabase_get_payload(department)
        if not raw:
            return default_payload()
        if isinstance(raw, str):
            raw = json.loads(raw)
        return migrate_payload(raw)

    conn = sqlite3.connect(DB_PATH)
    row = conn.execute(
        "SELECT payload_json FROM budget_entries WHERE department = ?", (department,)
    ).fetchone()
    conn.close()

    if not row:
        return default_payload()

    raw = json.loads(row[0])
    return migrate_payload(raw)


def load_all_payloads(departments: list[str]) -> dict:
    return {dept: load_payload(dept) for dept in departments}


def is_travel_section(section: str) -> bool:
    return "TRAVEL" in section.upper()


def item_cost_by_month(item: dict) -> list[float]:
    tb = item.get("travel_breakdown")
    if isinstance(tb, dict):
        # Travel subheads: monthly cost is sum of the breakdown rows.
        out: list[float] = []
        for m in MONTHS:
            total = 0.0
            for k in TRAVEL_COST_KEYS:
                total += _to_float(tb.get(k, {}).get(m, 0.0))
            out.append(total)
        return out
    if item.get("kind") == "amount":
        return [_to_float(item["amount"].get(m, 0.0)) for m in MONTHS]
    units = [_to_float(item["units"].get(m, 0.0)) for m in MONTHS]
    rate = [_to_float(item["rate"].get(m, 0.0)) for m in MONTHS]
    return [units[i] * rate[i] for i in range(len(MONTHS))]


def item_benefit_by_month(item: dict) -> list[float]:
    return [_to_float(item.get("benefit", {}).get(m, 0.0)) for m in MONTHS]


def item_fy_cost(item: dict) -> float:
    return float(sum(item_cost_by_month(item)))


def item_fy_benefit(item: dict) -> float:
    return float(sum(item_benefit_by_month(item)))


def item_fy_roi(item: dict) -> float | None:
    cost = item_fy_cost(item)
    if cost <= 0:
        return None
    return ((item_fy_benefit(item) - cost) / cost) * 100.0


def section_totals(payload: dict) -> dict:
    totals = {}
    for section in SECTION_NAMES:
        month_totals = {m: 0.0 for m in MONTHS}
        for item in payload.get("sections", {}).get(section, []):
            vals = item_cost_by_month(item)
            for i, month in enumerate(MONTHS):
                month_totals[month] += vals[i]
        totals[section] = {"months": month_totals, "fy": sum(month_totals.values())}
    return totals


def build_summary_dataframe(all_payloads: dict, departments: list[str]) -> pd.DataFrame:
    rows = []
    for section_idx, section in enumerate(SECTION_NAMES, start=1):
        sec_name = section.split(". ", 1)[1] if ". " in section else section
        rows.append({"#": str(section_idx), "Section / Period": sec_name})
        for period in MONTHS + ["FY Total"]:
            row = {"#": "", "Section / Period": f"  {period}"}
            dept_total = 0.0
            for dept in departments:
                sec = section_totals(all_payloads[dept])[section]
                val = sec["fy"] if period == "FY Total" else sec["months"][period]
                row[dept] = round(val, 2)
                dept_total += val
            row["GRAND TOTAL"] = round(dept_total, 2)
            rows.append(row)
        rows.append({"#": "", "Section / Period": ""})

    grand = {"#": "", "Section / Period": "TOTAL - ALL DEPARTMENTS"}
    for dept in departments:
        dept_grand = 0.0
        for section in SECTION_NAMES:
            dept_grand += section_totals(all_payloads[dept])[section]["fy"]
        grand[dept] = round(dept_grand, 2)
    grand["GRAND TOTAL"] = round(sum(grand[d] for d in departments), 2)
    rows.append(grand)

    return pd.DataFrame(rows)


def workbook_bytes(all_payloads: dict, departments: list[str], include_summary: bool) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    if include_summary:
        ws = wb.create_sheet("CONSOLIDATED SUMMARY")
        df_sum = build_summary_dataframe(all_payloads, departments)
        for col_idx, col_name in enumerate(df_sum.columns, start=1):
            c = ws.cell(row=1, column=col_idx, value=col_name)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F3864")
            c.alignment = Alignment(horizontal="center")
        for r_idx, (_, row) in enumerate(df_sum.iterrows(), start=2):
            for c_idx, col_name in enumerate(df_sum.columns, start=1):
                val = row[col_name]
                c = ws.cell(row=r_idx, column=c_idx, value=val)
                if isinstance(val, (int, float)):
                    c.number_format = '#,##0.00;(#,##0.00);"-"'
                if col_name == "Section / Period":
                    c.alignment = Alignment(horizontal="left")

    for dept in departments:
        payload = all_payloads[dept]
        ws = wb.create_sheet(dept[:31])
        headers = ["Section", "Subhead", "Description", "Metric"] + MONTHS + ["FY Total", "FY Benefit", "ROI %", "Support Docs", "Comment"]
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col_idx, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="2E5FAC")
            c.alignment = Alignment(horizontal="center")

        row_idx = 2
        for section in SECTION_NAMES:
            sec_name = section.split(". ", 1)[1] if ". " in section else section
            for item in payload.get("sections", {}).get(section, []):
                comment = item.get("comment", "")
                desc = item.get("description", "")
                is_travel = is_travel_section(section)
                fy_benefit = item_fy_benefit(item) if is_travel else 0.0
                fy_roi = item_fy_roi(item) if is_travel else None

                # Travel subheads: export breakdown rows + total.
                tb = item.get("travel_breakdown") if is_travel else None
                if is_travel and isinstance(tb, dict):
                    ensure_travel_breakdown(item)
                    breakdown_rows = []
                    for k in TRAVEL_COST_KEYS:
                        breakdown_rows.append((k, [_to_float(tb.get(k, {}).get(m, 0.0)) for m in MONTHS]))
                    breakdown_rows.append(("Total", item_cost_by_month(item)))

                    for metric_name, vals in breakdown_rows:
                        ws.cell(row=row_idx, column=1, value=sec_name)
                        ws.cell(row=row_idx, column=2, value=item.get("name", "Subhead"))
                        ws.cell(row=row_idx, column=3, value=desc)
                        ws.cell(row=row_idx, column=4, value=metric_name)
                        for i, v in enumerate(vals, start=5):
                            ws.cell(row=row_idx, column=i, value=round(v, 2)).number_format = '#,##0.00;(#,##0.00);"-"'
                        fy_total_col = 5 + len(MONTHS)
                        ws.cell(row=row_idx, column=fy_total_col, value=round(sum(vals), 2)).number_format = '#,##0.00;(#,##0.00);"-"'
                        if metric_name == "Total":
                            ws.cell(row=row_idx, column=fy_total_col + 1, value=round(fy_benefit, 2) if is_travel else "")
                            ws.cell(row=row_idx, column=fy_total_col + 2, value=round(fy_roi, 2) if fy_roi is not None else "")
                            ws.cell(row=row_idx, column=fy_total_col + 3, value="; ".join([a.get("name", "") for a in item.get("attachments", [])]))
                            ws.cell(row=row_idx, column=fy_total_col + 4, value=comment)
                        row_idx += 1
                    continue

                if item.get("kind") == "amount":
                    vals = item_cost_by_month(item)
                    ws.cell(row=row_idx, column=1, value=sec_name)
                    ws.cell(row=row_idx, column=2, value=item.get("name", "Subhead"))
                    ws.cell(row=row_idx, column=3, value=desc)
                    ws.cell(row=row_idx, column=4, value="Amount")
                    for i, v in enumerate(vals, start=5):
                        ws.cell(row=row_idx, column=i, value=round(v, 2)).number_format = '#,##0.00;(#,##0.00);"-"'
                    fy_total_col = 5 + len(MONTHS)
                    ws.cell(row=row_idx, column=fy_total_col, value=round(sum(vals), 2)).number_format = '#,##0.00;(#,##0.00);"-"'
                    ws.cell(row=row_idx, column=fy_total_col + 1, value=round(fy_benefit, 2) if is_travel else "")
                    ws.cell(row=row_idx, column=fy_total_col + 2, value=round(fy_roi, 2) if fy_roi is not None else "")
                    ws.cell(row=row_idx, column=fy_total_col + 3, value="; ".join([a.get("name", "") for a in item.get("attachments", [])]))
                    ws.cell(row=row_idx, column=fy_total_col + 4, value=comment)
                    row_idx += 1
                else:
                    units = [_to_float(item["units"].get(m, 0.0)) for m in MONTHS]
                    rate = [_to_float(item["rate"].get(m, 0.0)) for m in MONTHS]
                    value = [units[i] * rate[i] for i in range(len(MONTHS))]
                    for metric_name, vals in [("Units", units), ("Rate", rate), ("Value", value)]:
                        ws.cell(row=row_idx, column=1, value=sec_name)
                        ws.cell(row=row_idx, column=2, value=item.get("name", "Subhead"))
                        ws.cell(row=row_idx, column=3, value=desc)
                        ws.cell(row=row_idx, column=4, value=metric_name)
                        for i, v in enumerate(vals, start=5):
                            ws.cell(row=row_idx, column=i, value=round(v, 2)).number_format = '#,##0.00;(#,##0.00);"-"'
                        fy_total_col = 5 + len(MONTHS)
                        ws.cell(row=row_idx, column=fy_total_col, value=round(sum(vals), 2)).number_format = '#,##0.00;(#,##0.00);"-"'
                        if metric_name == "Value" and is_travel:
                            ws.cell(row=row_idx, column=fy_total_col + 1, value=round(fy_benefit, 2))
                            ws.cell(row=row_idx, column=fy_total_col + 2, value=round(fy_roi, 2) if fy_roi is not None else "")
                            ws.cell(row=row_idx, column=fy_total_col + 3, value="; ".join([a.get("name", "") for a in item.get("attachments", [])]))
                            ws.cell(row=row_idx, column=fy_total_col + 4, value=comment)
                        row_idx += 1

        # Column widths
        fy_total_col = 5 + len(MONTHS)
        support_docs_col = fy_total_col + 3
        comment_col = fy_total_col + 4

        for col in [1, 2, 3, 4, support_docs_col, comment_col]:
            if col in [1, 2]:
                ws.column_dimensions[chr(64 + col)].width = 24
            elif col == 3:
                ws.column_dimensions[chr(64 + col)].width = 28
            elif col == 4:
                ws.column_dimensions[chr(64 + col)].width = 16
            else:
                ws.column_dimensions[chr(64 + col)].width = 18

        for col_offset in range(5, support_docs_col):
            ws.column_dimensions[chr(64 + col_offset)].width = 12

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _editor_to_month_map(df: pd.DataFrame, row_name: str) -> dict:
    return {m: _to_float(df.loc[row_name, m]) for m in MONTHS}


def render_department_form(department: str, payload: dict, edit_locked: bool) -> dict:
    st.subheader(f"Department: {department}")
    if edit_locked:
        st.warning("Editing is locked by MASTER. You can view/download, but cannot change or save data.")
    else:
        st.caption("Users can add/edit subheads and choose calculation type. Travel subheads include ROI.")

    to_delete = []

    for section in SECTION_NAMES:
        with st.expander(section, expanded=False):
            section_items = payload["sections"].setdefault(section, [])

            add_col, _ = st.columns([1, 3])
            with add_col:
                if st.button("Add Subhead", key=f"add_{department}_{section}", disabled=edit_locked):
                    section_items.append(create_item("New Subhead", "units_rate"))
                    st.rerun()

            if not section_items:
                st.info("No subheads in this section. Click Add Subhead.")

            for idx, item in enumerate(section_items):
                item_id = item["id"]
                if is_travel_section(section):
                    item["kind"] = "travel_breakdown"
                    ensure_travel_breakdown(item)

                st.markdown(f"Subhead #{idx + 1}")
                c1, c2, c3 = st.columns([2, 1.2, 0.8])
                item["name"] = c1.text_input(
                    "Subhead Name",
                    value=item.get("name", "New Subhead"),
                    key=f"name_{department}_{item_id}",
                    disabled=edit_locked,
                )

                if is_travel_section(section):
                    c2.caption("Travel cost = Fare + Food + Lodging")
                else:
                    item["kind"] = c2.selectbox(
                        "Field Type",
                        options=["units_rate", "amount"],
                        format_func=lambda x: "Units x Rate" if x == "units_rate" else "Direct Amount",
                        index=0 if item.get("kind") == "units_rate" else 1,
                        key=f"kind_{department}_{item_id}",
                        disabled=edit_locked,
                    )
                if c3.button("Delete", key=f"del_{department}_{item_id}", disabled=edit_locked):
                    to_delete.append((section, item_id))

                item["description"] = st.text_area(
                    "Description",
                    value=item.get("description", ""),
                    key=f"desc_{department}_{item_id}",
                    disabled=edit_locked,
                    height=72,
                )

                if is_travel_section(section):
                    tb = item.get("travel_breakdown")
                    breakdown_df = pd.DataFrame(
                        [[_to_float(tb.get(k, {}).get(m, 0.0)) for m in MONTHS] for k in TRAVEL_COST_KEYS],
                        index=TRAVEL_COST_KEYS,
                        columns=MONTHS,
                    )

                    if edit_locked:
                        st.dataframe(breakdown_df, use_container_width=True)
                    else:
                        edited_tb = st.data_editor(
                            breakdown_df,
                            key=f"tb_{department}_{item_id}",
                            use_container_width=True,
                            num_rows="fixed",
                        )
                        for k in TRAVEL_COST_KEYS:
                            item["travel_breakdown"][k] = _editor_to_month_map(edited_tb, k)

                    calc_vals = item_cost_by_month(item)
                    st.dataframe(
                        pd.DataFrame([calc_vals], index=["Total (auto)"], columns=MONTHS),
                        use_container_width=True,
                    )
                else:
                    if item["kind"] == "amount":
                        amount_df = pd.DataFrame(
                            [[_to_float(item["amount"].get(m, 0.0)) for m in MONTHS]],
                            index=["Amount"],
                            columns=MONTHS,
                        )
                        if edit_locked:
                            st.dataframe(amount_df, use_container_width=True)
                        else:
                            edited_amount = st.data_editor(
                                amount_df,
                                key=f"amt_{department}_{item_id}",
                                use_container_width=True,
                                num_rows="fixed",
                            )
                            item["amount"] = _editor_to_month_map(edited_amount, "Amount")
                    else:
                        input_df = pd.DataFrame(
                            [
                                [_to_float(item["units"].get(m, 0.0)) for m in MONTHS],
                                [_to_float(item["rate"].get(m, 0.0)) for m in MONTHS],
                            ],
                            index=["Units", "Rate"],
                            columns=MONTHS,
                        )
                        if edit_locked:
                            st.dataframe(input_df, use_container_width=True)
                        else:
                            edited_ur = st.data_editor(
                                input_df,
                                key=f"ur_{department}_{item_id}",
                                use_container_width=True,
                                num_rows="fixed",
                            )
                            item["units"] = _editor_to_month_map(edited_ur, "Units")
                            item["rate"] = _editor_to_month_map(edited_ur, "Rate")
                        calc_vals = item_cost_by_month(item)
                        st.dataframe(
                            pd.DataFrame([calc_vals], index=["Value (auto)"], columns=MONTHS),
                            use_container_width=True,
                        )

                item["comment"] = st.text_input(
                    "Comment / Assumption",
                    value=item.get("comment", ""),
                    key=f"comment_{department}_{item_id}",
                    disabled=edit_locked,
                )

                if is_travel_section(section):
                    with st.expander("ROI & Supporting Docs", expanded=True):
                        benefit_df = pd.DataFrame(
                            [[_to_float(item["benefit"].get(m, 0.0)) for m in MONTHS]],
                            index=["Expected Return"],
                            columns=MONTHS,
                        )
                        if edit_locked:
                            st.dataframe(benefit_df, use_container_width=True)
                        else:
                            edited_benefit = st.data_editor(
                                benefit_df,
                                key=f"benefit_{department}_{item_id}",
                                use_container_width=True,
                                num_rows="fixed",
                            )
                            item["benefit"] = _editor_to_month_map(edited_benefit, "Expected Return")

                        fy_roi = item_fy_roi(item)
                        st.caption(
                            f"FY Cost: {item_fy_cost(item):,.2f} | FY Return: {item_fy_benefit(item):,.2f} | "
                            f"FY ROI: {(f'{fy_roi:.2f}%' if fy_roi is not None else 'N/A')}"
                        )

                        if not edit_locked:
                            uploaded = st.file_uploader(
                                "Attach supporting document",
                                type=["pdf", "xlsx", "docx", "png", "jpg", "jpeg"],
                                key=f"up_{department}_{item_id}",
                            )
                            err = add_attachment(item, uploaded)
                            if err:
                                st.warning(err)

                        docs = item.get("attachments", [])
                        if docs:
                            st.caption("Attached documents")
                            for doc in docs:
                                d1, d2 = st.columns([4, 1])
                                d1.download_button(
                                    f"Download {attachment_label(doc)}",
                                    data=base64.b64decode(doc.get("content_b64", "")),
                                    file_name=doc.get("name", "supporting_doc"),
                                    mime=doc.get("mime", "application/octet-stream"),
                                    key=f"dl_{department}_{item_id}_{doc.get('id')}",
                                )
                                if not edit_locked:
                                    if d2.button("Remove", key=f"rm_{department}_{item_id}_{doc.get('id')}"):
                                        item["attachments"] = [a for a in docs if a.get("id") != doc.get("id")]
                                        st.rerun()

                st.divider()

    for section, item_id in to_delete:
        payload["sections"][section] = [it for it in payload["sections"][section] if it["id"] != item_id]
    if to_delete:
        st.rerun()

    return payload


def build_travel_roi_report(payload: dict) -> pd.DataFrame:
    rows = []
    for section in SECTION_NAMES:
        if not is_travel_section(section):
            continue
        sec_label = section.split(". ", 1)[1] if ". " in section else section
        for item in payload.get("sections", {}).get(section, []):
            fy_roi = item_fy_roi(item)
            rows.append(
                {
                    "Section": sec_label,
                    "Subhead": item.get("name", "Subhead"),
                    "FY Cost": round(item_fy_cost(item), 2),
                    "FY Return": round(item_fy_benefit(item), 2),
                    "FY ROI %": round(fy_roi, 2) if fy_roi is not None else None,
                    "Docs": len(item.get("attachments", [])),
                }
            )
    return pd.DataFrame(rows)


def master_documents_rows(all_payloads: dict) -> list[dict]:
    rows = []
    for dept, payload in all_payloads.items():
        for section in SECTION_NAMES:
            if not is_travel_section(section):
                continue
            for item in payload.get("sections", {}).get(section, []):
                for att in item.get("attachments", []):
                    rows.append(
                        {
                            "Department": dept,
                            "Section": section,
                            "Subhead": item.get("name", "Subhead"),
                            "name": att.get("name", "supporting_doc"),
                            "mime": att.get("mime", "application/octet-stream"),
                            "size": int(_to_float(att.get("size", 0))),
                            "content_b64": att.get("content_b64", ""),
                            "id": att.get("id", uuid.uuid4().hex),
                        }
                    )
    return rows


def login_view(auth_map: dict, master_pw: str) -> None:
    render_header(compact=False)

    departments = login_domains(auth_map)
    dept_domains = load_department_domains()
    domain_options = sorted({
        *{
            department_domain_for_login(dept_domains, d)
            for d in departments
            if department_domain_for_login(dept_domains, d)
        },
        MASTER_DOMAIN_NAME,
    })
    domain_options = [d for d in domain_options if d]
    if not domain_options:
        domain_options = [MASTER_DOMAIN_NAME]

    outer = st.container()
    with outer:
        cols = st.columns([1, 1, 1])
        with cols[1]:
            try:
                card = st.container(border=True)
            except TypeError:
                card = st.container()
            with card:
                st.markdown("### Sign in")
                selected_domain = st.selectbox(
                    "Domain",
                    options=domain_options,
                    key="login_selected_domain",
                )
                last_domain = st.session_state.get("login_last_domain")
                if last_domain != selected_domain:
                    st.session_state["login_last_domain"] = selected_domain
                    st.session_state.pop("login_as", None)

                visible_departments = [
                    d
                    for d in departments
                    if department_domain_for_login(dept_domains, d) == selected_domain
                ]

                login_options = visible_departments
                if selected_domain == MASTER_DOMAIN_NAME:
                    login_options = ["MASTER"] + login_options

                if not login_options:
                    st.warning("No departments found under this domain.")
                    return

                login_as = st.selectbox(
                    "Login As",
                    options=login_options,
                    key="login_as",
                )
                pw = st.text_input("Password", type="password", key="login_pw")
                submit = st.button("Login", use_container_width=True)

    if not submit:
        return

    if login_as == "MASTER" and pw == master_pw:
        st.session_state.authenticated = True
        st.session_state.role = "master"
        st.session_state.department = None
        st.rerun()

    if login_as != "MASTER" and check_domain_password(auth_map, login_as, pw):
        st.session_state.authenticated = True
        st.session_state.role = "department"
        st.session_state.department_domain = login_as
        if login_as == SUPPLY_CHAIN_DOMAIN and supply_chain_enabled(auth_map):
            st.session_state.supply_chain_segment = SUPPLY_CHAIN_SEGMENTS[0]
            st.session_state.department = st.session_state.supply_chain_segment
        else:
            st.session_state.department = login_as
        st.rerun()

    st.error("Invalid password")


def app_view(auth_map: dict, master_pw: str) -> None:
    all_departments = sorted(
        {
            *[d for d in auth_map.keys() if d != SUPPLY_CHAIN_DOMAIN],
            *SUPPLY_CHAIN_SEGMENTS,
        }
    )

    settings = load_app_settings()
    edit_locked = bool(settings.get("edit_locked", False))
    view_locked = bool(settings.get("view_locked", False))

    with st.sidebar:
        st.write(f"Role: {st.session_state.role}")
        if st.session_state.role == "department":
            domain = st.session_state.get("department_domain") or st.session_state.department
            st.write(f"Department: {domain}")

            if domain == SUPPLY_CHAIN_DOMAIN and supply_chain_enabled(auth_map):
                seg = st.selectbox(
                    "Segregation",
                    options=SUPPLY_CHAIN_SEGMENTS,
                    index=SUPPLY_CHAIN_SEGMENTS.index(st.session_state.get("supply_chain_segment", SUPPLY_CHAIN_SEGMENTS[0]))
                    if st.session_state.get("supply_chain_segment") in SUPPLY_CHAIN_SEGMENTS
                    else 0,
                )
                if seg != st.session_state.get("supply_chain_segment"):
                    st.session_state.supply_chain_segment = seg
                    st.session_state.department = seg
                    st.rerun()

                st.caption(f"Active: {st.session_state.department}")
            if edit_locked:
                st.warning("Editing is locked")
            if view_locked:
                st.error("Viewing is locked")

        if st.session_state.role == "master":
            new_view_locked = st.toggle(
                "Block department viewing",
                value=view_locked,
                help="When ON, departments cannot view any data or download files.",
            )
            if new_view_locked != view_locked:
                save_app_settings(
                    {
                        **settings,
                        "view_locked": bool(new_view_locked),
                        "locked_by": "MASTER",
                        "locked_at": datetime.now().isoformat(timespec="seconds") if new_view_locked else settings.get("locked_at"),
                    }
                )
                st.rerun()

            new_locked = st.toggle(
                "Lock editing for departments",
                value=edit_locked,
                help="When ON, departments can view/download but cannot add, delete, edit, upload, or save data.",
            )
            if new_locked != edit_locked:
                save_app_settings(
                    {
                        **settings,
                        "edit_locked": bool(new_locked),
                        "locked_by": "MASTER",
                        "locked_at": datetime.now().isoformat(timespec="seconds") if new_locked else None,
                    }
                )
                st.rerun()
        if st.button("Logout"):
            st.session_state.authenticated = False
            st.session_state.role = None
            st.session_state.department = None
            st.session_state.department_domain = None
            st.session_state.supply_chain_segment = None
            st.rerun()

    render_header(compact=True)

    if st.session_state.role == "department":
        dept = st.session_state.department

        if view_locked:
            st.title(APP_TITLE)
            st.error("Access is temporarily disabled by MASTER. Please try again later.")
            return

        work_key = f"work_payload_{dept}"
        if work_key not in st.session_state:
            st.session_state[work_key] = load_payload(dept)

        current_payload = st.session_state[work_key]
        current_payload = render_department_form(dept, current_payload, edit_locked=edit_locked)
        st.session_state[work_key] = current_payload

        st.subheader("Travel ROI Snapshot")
        roi_df = build_travel_roi_report(current_payload)
        if not roi_df.empty:
            st.dataframe(roi_df, use_container_width=True)
        else:
            st.info("No travel subheads found yet.")

        col1, col2 = st.columns([1, 1])
        with col1:
            if st.button("Save Department Data", type="primary", disabled=edit_locked):
                if edit_locked:
                    st.warning("Editing is locked. MASTER must unlock before you can save changes.")
                else:
                    save_payload(dept, current_payload)
                    st.success("Department data saved.")

        with col2:
            xlsx = workbook_bytes({dept: current_payload}, [dept], include_summary=False)
            st.download_button(
                "Download My Department Excel",
                data=xlsx,
                file_name=f"Budget_{dept.replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        render_shared_sheets_panel(edit_locked=edit_locked, view_locked=view_locked)

    if st.session_state.role == "master":
        all_payloads = load_all_payloads(all_departments)

        status_rows = []
        for dept in all_departments:
            upd = all_payloads[dept].get("updated_at")
            status_rows.append({"Department": dept, "Last Updated": upd or "Not submitted"})

        st.subheader("Submission Status")
        st.dataframe(pd.DataFrame(status_rows), use_container_width=True)

        st.subheader("Detailed Consolidated Summary")
        df_sum = build_summary_dataframe(all_payloads, all_departments)
        st.dataframe(df_sum, use_container_width=True, height=600)

        st.subheader("Travel ROI Summary")
        travel_rows = []
        for dept in all_departments:
            roi_df = build_travel_roi_report(all_payloads[dept])
            if roi_df.empty:
                continue
            roi_df.insert(0, "Department", dept)
            travel_rows.append(roi_df)
        if travel_rows:
            st.dataframe(pd.concat(travel_rows, ignore_index=True), use_container_width=True)
        else:
            st.info("No travel ROI data submitted yet.")

        st.subheader("Travel Supporting Documents")
        doc_rows = master_documents_rows(all_payloads)
        if doc_rows:
            doc_df = pd.DataFrame(
                [{
                    "Department": r["Department"],
                    "Section": r["Section"],
                    "Subhead": r["Subhead"],
                    "Document": r["name"],
                    "Size (KB)": int((r["size"] + 1023) // 1024),
                } for r in doc_rows]
            )
            st.dataframe(doc_df, use_container_width=True)
            for r in doc_rows:
                st.download_button(
                    f"Download {r['Department']} - {r['Subhead']} - {r['name']}",
                    data=base64.b64decode(r["content_b64"]),
                    file_name=r["name"],
                    mime=r["mime"],
                    key=f"mdoc_{r['id']}",
                )
        else:
            st.info("No supporting documents uploaded yet.")

        xlsx = workbook_bytes(all_payloads, all_departments, include_summary=True)
        st.download_button(
            "Download Master Consolidated Excel",
            data=xlsx,
            file_name="Budget_Master_Consolidated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        render_shared_sheets_panel(edit_locked=False, view_locked=False)


def main() -> None:
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    ensure_db()
    auth_map, master_pw = load_auth_map()

    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
        st.session_state.role = None
        st.session_state.department = None
        st.session_state.department_domain = None
        st.session_state.supply_chain_segment = None

    if not st.session_state.authenticated:
        login_view(auth_map, master_pw)
        return

    app_view(auth_map, master_pw)


if __name__ == "__main__":
    main()
